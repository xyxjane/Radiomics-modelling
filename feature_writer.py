'''
Description:
Author: xueyuxin_jane
Date: 2020-08-18 16:55:54
LastEditTime: 2021-05-24 19:19:58
LastEditors: xueyuxin_jane
'''
import os
import pandas as pd
from feature_extraction import feature_extraction
# from read_data import find_minus_scan
import yaml
import copy
from openpyxl import load_workbook
import openpyxl


# PET_CT_path = "PSMA_75"
data_path = "FDG_PSMA"
fdg_excel_path = "FDG_features_LNI.xlsx"
psma_excel_path = "PSMA_features_75_LNI.xlsx"
parafile = "LungNode_Pre/Params.yaml"
test_excel_path = "test.xlsx"

def append_df_to_excel(filename, df, sheet_name='pet_features_bin128', startrow=None, startcol=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
   # from openpyxl import load_workbook
   # import pandas as pd
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    df_old = pd.read_excel(filename)
    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        if startcol is None and sheet_name in writer.book.sheetnames:
            startcol = df_old.shape[1]
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    df.to_excel(writer, sheet_name, startrow=startrow,
                startcol=startcol, **to_excel_kwargs)
    # save the workbook
    writer.save()

def save_data(df, excel_path, sheet):
    '''
    description: save new df_data to new sheet and resume previous sheet
    param :
    df:new data needs to be saved
    excel_path: excel_path(can be empty or not)
    sheet: sheet_name of df to be save in
    return {None}
    '''
    # 写入数据 encoding="utf-8-sig" 看情况而用哦
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    book = openpyxl.load_workbook(writer.path)
    writer.book = book
    df.to_excel(excel_writer=writer, sheet_name=sheet, index=None)
    writer.save()
    writer.close()

def main(data_path, parafile_path, excel_path):
    all_patients = os.listdir(data_path)
    patient_num = 0 
    all_feature_dic = {}
    for patient in all_patients:
        print(patient)
        if "DS_Store" not in patient:
            patient_path = os.path.join(data_path,patient)
            patient_PETs = os.listdir(patient_path)
            for PET_type in patient_PETs:
                if "DS_Store" not in PET_type:
                    if "PSMA" in PET_type:
                        pass
                    else:
                        FDG_file = PET_type
                        PET_CT_path = os.path.join(patient_path,FDG_file)
                        ct_data_path = os.path.join(PET_CT_path, "PET.mha")
                        print(ct_data_path)
                        # print(patient)
                        ct_label_path = os.path.join(PET_CT_path, "label.mha")
                        featureVector_dic = feature_extraction(
                            ct_data_path, ct_label_path, parafile_path, image_type="Original")
                        # print(featureVector_dic)
                        for featurelistName in featureVector_dic.keys():
                            featureVector = featureVector_dic[featurelistName]
                            featureVector_copy = featureVector.copy()
                            for featureName in featureVector_copy:
                                new_featureName = "pet_"+"original_"+featurelistName+"_"+featureName
                                featureVector[new_featureName] = featureVector.pop(featureName)
                                if patient_num == 0:
                                    all_feature_dic[new_featureName] = [
                                        featureVector[new_featureName]]
                                    # print("writing unexisting",all_feature_dic)
                                else:
                                    print(all_feature_dic[new_featureName])
                                    all_feature_dic[new_featureName].append(
                                        featureVector[new_featureName])

    # for filterName in featureVector_dic.keys():
    #   featurelists = featureVector_dic[filterName]
    #   for featurelistName in featurelists.keys():
    #     featureVector = featurelists[featurelistName]
    #     featureVector_copy = featureVector.copy()
    #     current_feature_dic = {}
    #     for featureName in featureVector_copy.keys():
    #       new_featureName = "pet_"+filterName+featurelistName+"_"+featureName
    #       featureVector[new_featureName] = featureVector.pop(featureName)
    #       if patient_num == 0:
    #         all_feature_dic[new_featureName] = [featureVector[new_featureName]]
    #         # print("writing unexisting",single_feature_dic)
    #       else:
    #         all_feature_dic[new_featureName].append(featureVector[new_featureName])
            for featureName in all_feature_dic.keys():
                print(featureName,len(all_feature_dic[featureName]))
            patient_num = patient_num + 1
            print("writing feature vector")
    # print(origin_feature_vector)
    df = pd.DataFrame(all_feature_dic)
    if patient_num == 0:
        append_df_to_excel(excel_path, df, startrow=0, startcol=0, index=False)
    else:
        append_df_to_excel(excel_path, df, startrow=0,index=False)

def save_FDG_patient_data(FDG_data_path, PSMA_featurea_excel_path, FDG_features_excel_path):
    all_FDG_patients = os.listdir(FDG_data_path)
    all_list = list(range(1,75,1))
    for patient in all_FDG_patients:
        all_list.remove(int(patient))
    feature_df = pd.read_excel(PSMA_featurea_excel_path, sheet_name="PET_rfe_35x2")
    feature_copy = feature_df.copy(deep=True)
    selected_feature_df = feature_df.drop(feature_copy.index[[x-1 for x in all_list]])
    save_data(selected_feature_df,FDG_features_excel_path,"PSMA_final_features")



if __name__ == "__main__":
    # main(data_path, parafile, excel_path) 
    # main(PET_CT_path,parafile)
    # df1 = pd.DataFrame({'One':[1,2,3]})
    # df2 = pd.DataFrame({"Two":[6,8,9]})
    # append_df_to_excel(test_excel_path,df2,startrow=0,startcol=0,index=False)
    save_FDG_patient_data(data_path,psma_excel_path,fdg_excel_path)